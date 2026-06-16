import os
import io
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from backend.database import SessionLocal, init_db, User, Patient, HealthRecord, TriageRecord, Appointment, PatientPrescription, hash_password
from backend.medication_service import analyze_prescription
import json
from backend.triage import classify_triage
from backend.ml_model import predict_disease_risk
from backend.rag_engine import generate_treatment_recommendations, build_vector_store

# Initialize DB on startup
init_db()

# Pre-build vector store if guidelines exist
try:
    build_vector_store()
except Exception as e:
    print(f"Error pre-building vector store: {e}")

app = FastAPI(title="HealthAI - Intelligent Healthcare Support API")

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# DB Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Pydantic schemas
class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    role: str
    full_name: str

class UserResponse(BaseModel):
    id: int
    username: str
    role: str
    full_name: str

class PatientCreate(BaseModel):
    name: str
    age: int
    gender: str
    contact: str
    medical_history: str = ""
    allergies: str = ""

class HealthRecordCreate(BaseModel):
    symptoms: str
    systolic_bp: int
    diastolic_bp: int
    heart_rate: int
    temperature: float
    lab_results: str = ""
    diagnosis: str = ""
    prescription: str = ""
    notes: str = ""
    lab_report_image: Optional[str] = None

class TriageRequest(BaseModel):
    symptoms: str
    age: int
    systolic_bp: Optional[int] = None
    diastolic_bp: Optional[int] = None
    heart_rate: Optional[int] = None
    temperature: Optional[float] = None
    lab_results: Optional[str] = ""

class PredictionRequest(BaseModel):
    age: int
    gender: str
    systolic_bp: int
    diastolic_bp: int
    heart_rate: int
    temperature: float
    fasting_blood_sugar: int
    cholesterol: int
    hemoglobin: float
    has_cough: bool
    has_chest_pain: bool
    has_dyspnea: bool
    has_fatigue: bool

class RecommendationRequest(BaseModel):
    diagnosis: str
    symptoms: str
    medical_history: str = ""

class AppointmentCreate(BaseModel):
    patient_id: int
    doctor_id: Optional[int] = None
    appointment_date: str # ISO string format
    reason: str

class MedicationAnalyzeRequest(BaseModel):
    prescription_text: str
    allergies: Optional[str] = ""

class PrescriptionSaveRequest(BaseModel):
    prescription_text: str
    prescription_image: Optional[str] = None


# Authentication Routes
@app.post("/auth/login", response_model=UserResponse)
def login(req: LoginRequest, db: Session = Depends(get_db)):
    pwd_hash = hash_password(req.password)
    user = db.query(User).filter(User.username == req.username, User.password_hash == pwd_hash).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid username or password")
    return UserResponse(id=user.id, username=user.username, role=user.role, full_name=user.full_name)

@app.post("/auth/register", response_model=UserResponse)
def register(req: RegisterRequest, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.username == req.username).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Username already exists")
    
    if req.role not in ['Patient', 'Doctor', 'Nurse', 'Administrator', 'Super Admin']:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid user role")
        
    new_user = User(
        username=req.username,
        password_hash=hash_password(req.password),
        role=req.role,
        full_name=req.full_name
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return UserResponse(id=new_user.id, username=new_user.username, role=new_user.role, full_name=new_user.full_name)

# Patients Management Routes
@app.get("/patients")
def list_patients(search: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Patient)
    if search:
        query = query.filter(Patient.name.like(f"%{search}%") | Patient.id.like(f"%{search}%"))
    return query.all()

@app.post("/patients")
def create_patient(req: PatientCreate, db: Session = Depends(get_db)):
    patient = Patient(**req.dict())
    db.add(patient)
    db.commit()
    db.refresh(patient)
    return patient

@app.get("/patients/{patient_id}")
def get_patient(patient_id: int, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient

# Health Record / Visit History
@app.get("/patients/{patient_id}/records")
def get_health_records(patient_id: int, db: Session = Depends(get_db)):
    records = db.query(HealthRecord).filter(HealthRecord.patient_id == patient_id).order_by(HealthRecord.visit_date.desc()).all()
    return records

@app.post("/patients/{patient_id}/records")
def create_health_record(patient_id: int, req: HealthRecordCreate, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
        
    record = HealthRecord(patient_id=patient_id, **req.dict(), visit_date=datetime.utcnow())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record

# AI & Decision Support Modules
@app.post("/predict")
def predict_risk(req: PredictionRequest):
    try:
        res = predict_disease_risk(**req.dict())
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Prediction error: {str(e)}")

@app.post("/triage")
def triage_patient(req: TriageRequest):
    try:
        res = classify_triage(**req.dict())
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Triage error: {str(e)}")

@app.get("/triage/records")
def list_triage_records(db: Session = Depends(get_db)):
    records = db.query(TriageRecord).order_by(TriageRecord.created_at.desc()).all()
    # Join patient names
    res = []
    for r in records:
        res.append({
            "id": r.id,
            "patient_id": r.patient_id,
            "patient_name": r.patient.name if r.patient else "Unknown",
            "patient_age": r.patient.age if r.patient else 0,
            "priority_level": r.priority_level,
            "symptom_severity": r.symptom_severity,
            "recommended_department": r.recommended_department,
            "status": r.status,
            "created_at": r.created_at
        })
    return res

@app.post("/triage/submit")
def submit_triage_record(patient_id: int, req: TriageRequest, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
        
    triage_info = classify_triage(**req.dict())
    
    triage_rec = TriageRecord(
        patient_id=patient_id,
        priority_level=triage_info["priority_level"],
        symptom_severity=req.symptoms,
        recommended_department=triage_info["recommended_department"],
        status="Pending"
    )
    db.add(triage_rec)
    db.commit()
    db.refresh(triage_rec)
    return {
        "id": triage_rec.id,
        "triage_result": triage_info
    }

@app.post("/triage/{triage_id}/resolve")
def resolve_triage(triage_id: int, status_val: str = "Checked", db: Session = Depends(get_db)):
    rec = db.query(TriageRecord).filter(TriageRecord.id == triage_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Triage record not found")
    rec.status = status_val
    db.commit()
    return {"status": "success"}

@app.post("/recommendations")
def get_recommendations(req: RecommendationRequest):
    try:
        res = generate_treatment_recommendations(**req.dict())
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"RAG recommendation error: {str(e)}")

# Appointments Routes
@app.get("/appointments")
def list_appointments(db: Session = Depends(get_db)):
    appointments = db.query(Appointment).order_by(Appointment.appointment_date.asc()).all()
    res = []
    for a in appointments:
        doc = db.query(User).filter(User.id == a.doctor_id).first() if a.doctor_id else None
        res.append({
            "id": a.id,
            "patient_id": a.patient_id,
            "patient_name": a.patient.name if a.patient else "Unknown",
            "patient_age": a.patient.age if a.patient else 0,
            "patient_gender": a.patient.gender if a.patient else "Unknown",
            "patient_contact": a.patient.contact if a.patient else "Unknown",
            "patient_history": a.patient.medical_history if a.patient else "",
            "patient_allergies": a.patient.allergies if a.patient else "",
            "doctor_name": doc.full_name if doc else "Unassigned",
            "doctor_role": doc.role if doc else "Unassigned",
            "appointment_date": a.appointment_date.strftime("%Y-%m-%d %H:%M") if hasattr(a.appointment_date, "strftime") else str(a.appointment_date),
            "reason": a.reason,
            "status": a.status
        })
    return res


@app.post("/appointments")
def create_appointment(req: AppointmentCreate, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == req.patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
        
    target_date = datetime.fromisoformat(req.appointment_date)
    
    # 1. Doctor conflict check
    if req.doctor_id:
        conflict_doc = db.query(Appointment).filter(
            Appointment.doctor_id == req.doctor_id,
            Appointment.appointment_date == target_date,
            Appointment.status == "Scheduled"
        ).first()
        if conflict_doc:
            raise HTTPException(
                status_code=400,
                detail="The selected doctor already has a scheduled appointment at this exact date and time."
            )
            
    # 2. Patient conflict check
    conflict_pat = db.query(Appointment).filter(
        Appointment.patient_id == req.patient_id,
        Appointment.appointment_date == target_date,
        Appointment.status == "Scheduled"
    ).first()
    if conflict_pat:
        raise HTTPException(
            status_code=400,
            detail="The patient already has a scheduled appointment at this exact date and time."
        )

    appt = Appointment(
        patient_id=req.patient_id,
        doctor_id=req.doctor_id,
        appointment_date=target_date,
        reason=req.reason,
        status="Scheduled"
    )
    db.add(appt)
    db.commit()
    db.refresh(appt)
    return appt


@app.patch("/appointments/{appointment_id}")
def update_appointment_status(appointment_id: int, status_val: str, db: Session = Depends(get_db)):
    appt = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not appt:
        raise HTTPException(status_code=404, detail="Appointment not found")
    appt.status = status_val
    db.commit()
    return {"status": "success"}


# Analytics Dashboard Data & Reporting Endpoints
@app.get("/analytics/dashboard")
def get_analytics_dashboard_data(db: Session = Depends(get_db)):
    total_patients = db.query(Patient).count()
    total_records = db.query(HealthRecord).count()
    total_appointments = db.query(Appointment).count()
    
    # Disease Trends (Top diagnoses)
    records = db.query(HealthRecord.diagnosis).all()
    diagnoses_counts = {}
    for r in records:
        if r.diagnosis:
            diag = r.diagnosis.split('(')[0].strip() # Clean name e.g. "Stage 1 Hypertension (improved)" -> "Stage 1 Hypertension"
            diagnoses_counts[diag] = diagnoses_counts.get(diag, 0) + 1
            
    disease_trends = [{"disease": k, "count": v} for k, v in diagnoses_counts.items()]
    disease_trends = sorted(disease_trends, key=lambda x: x["count"], reverse=True)[:5]
    
    # Triage distribution
    triage_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    triages = db.query(TriageRecord.priority_level).all()
    for t in triages:
        if t.priority_level in triage_counts:
            triage_counts[t.priority_level] += 1
            
    triage_data = [{"priority": k, "count": v} for k, v in triage_counts.items()]
    
    # System Alerts: Patients requiring immediate attention (Critical/High and unresolved)
    alert_records = db.query(TriageRecord).filter(
        TriageRecord.priority_level.in_(["Critical", "High"]),
        TriageRecord.status == "Pending"
    ).all()
    
    alerts = []
    for r in alert_records:
        alerts.append({
            "patient_name": r.patient.name if r.patient else "Unknown",
            "priority": r.priority_level,
            "symptom_severity": r.symptom_severity,
            "department": r.recommended_department,
            "time_elapsed": "Unresolved"
        })
        
    return {
        "metrics": {
            "total_patients": total_patients,
            "total_records": total_records,
            "total_appointments": total_appointments
        },
        "disease_trends": disease_trends,
        "triage_distribution": triage_data,
        "alerts": alerts
    }

@app.get("/analytics/export/pdf")
def export_pdf_report(db: Session = Depends(get_db)):
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="FPDF not installed")
        
    data = get_analytics_dashboard_data(db)
    metrics = data["metrics"]
    disease_trends = data["disease_trends"]
    triage_dist = data["triage_distribution"]
    alerts = data["alerts"]
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", "B", 18)
    pdf.cell(0, 10, "HealthAI Clinical Operations & Analytics Report", ln=True, align="C")
    pdf.set_font("helvetica", "I", 10)
    pdf.cell(0, 10, f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", ln=True, align="C")
    pdf.ln(10)
    
    # Metrics
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, "1. Executive Summary Metrics", ln=True)
    pdf.set_font("helvetica", "", 11)
    pdf.cell(0, 8, f"Total Managed Patients: {metrics['total_patients']}", ln=True)
    pdf.cell(0, 8, f"Total Clinical Encounters (Health Records): {metrics['total_records']}", ln=True)
    pdf.cell(0, 8, f"Total Appointments Scheduled: {metrics['total_appointments']}", ln=True)
    pdf.ln(5)
    
    # Disease trends
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, "2. Top Diagnosed Diseases & Trends", ln=True)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(120, 8, "Diagnosed Condition", border=1)
    pdf.cell(40, 8, "Encounters", border=1, ln=True)
    pdf.set_font("helvetica", "", 10)
    for trend in disease_trends:
        pdf.cell(120, 8, trend["disease"], border=1)
        pdf.cell(40, 8, str(trend["count"]), border=1, ln=True)
    pdf.ln(5)
    
    # Triage
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, "3. Patient Triage Distribution", ln=True)
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(80, 8, "Priority Level", border=1)
    pdf.cell(80, 8, "Count", border=1, ln=True)
    pdf.set_font("helvetica", "", 10)
    for td in triage_dist:
        pdf.cell(80, 8, td["priority"], border=1)
        pdf.cell(80, 8, str(td["count"]), border=1, ln=True)
    pdf.ln(5)
    
    # Alerts
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(0, 10, "4. High-Risk Pending Alerts", ln=True)
    if not alerts:
        pdf.set_font("helvetica", "I", 10)
        pdf.cell(0, 8, "No active critical or high-risk pending triage alerts.", ln=True)
    else:
        pdf.set_font("helvetica", "B", 10)
        pdf.cell(50, 8, "Patient Name", border=1)
        pdf.cell(30, 8, "Priority", border=1)
        pdf.cell(50, 8, "Symptoms", border=1)
        pdf.cell(50, 8, "Dept", border=1, ln=True)
        pdf.set_font("helvetica", "", 8)
        for alert in alerts:
            # truncate text for cell
            name = alert["patient_name"][:20]
            severity = alert["symptom_severity"][:25]
            dept = alert["department"][:25]
            pdf.cell(50, 8, name, border=1)
            pdf.cell(30, 8, alert["priority"], border=1)
            pdf.cell(50, 8, severity, border=1)
            pdf.cell(50, 8, dept, border=1, ln=True)
            
    # Save to byte stream
    pdf_bytes = pdf.output(dest='S')
    
    headers = {
        'Content-Disposition': 'attachment; filename="healthai_analytics_report.pdf"'
    }
    return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf", headers=headers)

@app.get("/analytics/export/excel")
def export_excel_report(db: Session = Depends(get_db)):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
        
    data = get_analytics_dashboard_data(db)
    
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws1 = wb.active
    ws1.title = "Summary Metrics"
    ws1.append(["HealthAI Analytics Report Summary"])
    ws1.append(["Generated on", datetime.utcnow().isoformat()])
    ws1.append([])
    ws1.append(["Metric", "Value"])
    ws1.append(["Total Managed Patients", data["metrics"]["total_patients"]])
    ws1.append(["Total Health Records", data["metrics"]["total_records"]])
    ws1.append(["Total Appointments", data["metrics"]["total_appointments"]])
    
    # Disease Trends Sheet
    ws2 = wb.create_sheet(title="Disease Trends")
    ws2.append(["Condition/Disease", "Frequency Count"])
    for trend in data["disease_trends"]:
        ws2.append([trend["disease"], trend["count"]])
        
    # Triage Sheet
    ws3 = wb.create_sheet(title="Triage Distribution")
    ws3.append(["Priority Level", "Patient Count"])
    for td in data["triage_distribution"]:
        ws3.append([td["priority"], td["count"]])
        
    # Alerts Sheet
    ws4 = wb.create_sheet(title="Active Alerts")
    ws4.append(["Patient Name", "Priority Level", "Symptom Severity", "Recommended Department"])
    for alert in data["alerts"]:
        ws4.append([alert["patient_name"], alert["priority"], alert["symptom_severity"], alert["department"]])
        
    # Save to memory stream
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="healthai_analytics_report.xlsx"'
    }
    return StreamingResponse(
        excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# Medication Safety Routes
@app.post("/medications/analyze")
def analyze_medications(req: MedicationAnalyzeRequest):
    return analyze_prescription(req.prescription_text, allergies=req.allergies)


@app.get("/patients/{patient_id}/prescriptions")
def get_patient_prescriptions(patient_id: int, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    records = db.query(PatientPrescription).filter(
        PatientPrescription.patient_id == patient_id
    ).order_by(PatientPrescription.created_at.desc()).all()
    return records


@app.post("/patients/{patient_id}/prescriptions")
def save_patient_prescription(patient_id: int, req: PrescriptionSaveRequest, db: Session = Depends(get_db)):
    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    analysis = analyze_prescription(req.prescription_text, allergies=patient.allergies or "")
    presc = PatientPrescription(
        patient_id=patient_id,
        prescription_text=req.prescription_text,
        prescription_image=req.prescription_image,
        drugs_detected=json.dumps(analysis["detected_drugs"]),
        analysis_json=json.dumps(analysis),
    )
    db.add(presc)
    db.commit()
    db.refresh(presc)
    return {"prescription": presc, "analysis": analysis}
